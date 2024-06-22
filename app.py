from typing import Tuple
import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd   
import pathlib
from openpyxl import Workbook

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.apparence()
        self.todo_sistema()

    def layout_config(self):
        self.title("Sistema de Gestao")
        self.geometry("700x500")

    def apparence(self):
        self.lb_apm = ctk.CTkLabel(self,text="Tema", bg_color="transparent", text_color=['#000','#fff']).place(x=50,y=430)
        self.opt_apm = ctk.CTkOptionMenu(self,values=["Light","Dark", "System"],    command=self.change_apm).place(x=50,y=460)

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=700,height=50,corner_radius=0,bg_color="teal",fg_color="teal")
        frame.place(x=0,y=10)
        title = ctk.CTkLabel(frame,text = "Sistema de gestao",font=("Century Gothic bold",24), text_color="#fff").place(x=190,y=10)

        span= ctk.CTkLabel(frame,text = "Preencha todos os campos do formulario",font=("Century Gothic bold",16), text_color=["#000","#fff"]).place(x=50,y=70)
        
        ficheiro = pathlib.Path("Subs.xlsx")
            
        if ficheiro.exists():
            pass
        else:
                ficheiro=Workbook()
                folha = ficheiro.active
                folha['A1']= "Nome"
                folha['B1']="genero"
                folha['C1']="userX"
                folha['D1']="subs"
                folha['E1']="tempo"
                folha['F1']="numero"

                ficheiro.save("subs.xlsx")
        def submit():
            nome = nome_value.get()
            genero = genero_value.get()
            userX = userX_value.get()
            subs = subs_value.get()
            tempo = tempo_value.get()
            numero = numero_value.get()
   
    
            ficheiro = openpyxl.load_workbook('Subs.xlsx')
            folha = ficheiro.active
            folha.cell(column=1,row=folha.max_row+1, value=nome)
            folha.cell(column=2,row=folha.max_row, value=genero)
            folha.cell(column=3,row=folha.max_row, value=userX)
            folha.cell(column=4,row=folha.max_row, value=subs)
            folha.cell(column=5,row=folha.max_row, value=tempo)
            folha.cell(column=6,row=folha.max_row, value=numero)

            ficheiro.save(r"Subs.xlsx")
            messagebox.showinfo("Sistema", "Dados salvos com sucesso")

            
        def clear():
            nome_value.set("")
            genero_value.set("")
            userX_value.set("")
            subs_value.set("")
            tempo_value.set("")
            numero_value.set("")


          #text variables  

        nome_value = StringVar()
        genero_value = StringVar()
        userX_value = StringVar()
        subs_value = StringVar()
        tempo_value = StringVar()
        numero_value = StringVar()

        #entrys
        entry_nome = ctk.CTkEntry(self, width=220, textvariable=nome_value,font=("Century Gothic bold",16),fg_color="transparent")
        entry_genero = ctk.CTkEntry(self, width=220,textvariable=genero_value, font=("Century Gothic bold",16),fg_color="transparent")
        entry_UserX= ctk.CTkEntry(self, width=220,textvariable=userX_value, font=("Century Gothic bold",16),fg_color="transparent")
        entry_Subs = ctk.CTkEntry(self, width=220,textvariable=subs_value,font=("Century Gothic bold",16),fg_color="transparent")
        entry_TempoSub = ctk.CTkEntry(self, width=220,textvariable=tempo_value, font=("Century Gothic bold",16),fg_color="transparent")
        entry_Numero = ctk.CTkEntry(self, width=220,textvariable=numero_value,font=("Century Gothic bold",16),fg_color="transparent")

        #labels
        lb_nome= ctk.CTkLabel(self,text = "Nome",font=("Century Gothic bold",16), text_color=["#000","#fff"])
        lb_genero= ctk.CTkLabel(self,text = "Genero",font=("Century Gothic bold",16), text_color=["#000","#fff"])
        lb_UserX= ctk.CTkLabel(self,text = "User",font=("Century Gothic bold",16), text_color=["#000","#fff"])
        lb_Subs= ctk.CTkLabel(self,text = "Subs",font=("Century Gothic bold",16), text_color=["#000","#fff"])
        lb_TempoSub= ctk.CTkLabel(self,text = "Tempo de sub",font=("Century Gothic bold",16), text_color=["#000","#fff"])
        lb_Numero= ctk.CTkLabel(self,text = "Numero",font=("Century Gothic bold",16), text_color=["#000","#fff"])

        #buttons
        btn = ctk.CTkButton(self,text="Salvar Dados".upper(),command=submit,fg_color='#151',hover_color='#131').place(x=300,y=420)
        btn = ctk.CTkButton(self,text="Limpar Campos".upper(),command=clear,fg_color='#555',hover_color='#333').place(x=500,y=420)
        

    #posição dos elementos na janela
        lb_nome.place(x=50, y=120)
        entry_nome.place(x=50, y=150)

        lb_genero.place(x=450, y=120)
        entry_genero.place(x=450, y=150)

        lb_UserX.place(x=450,y=190)
        entry_UserX.place(x=450,y=220)

        lb_Subs.place(x=450, y=260)
        entry_Subs.place(x=450, y=300)

        lb_TempoSub.place(x=50, y=190)
        entry_TempoSub.place(x=50, y=220)

        lb_Numero.place(x=50, y=260)
        entry_Numero.place(x=50, y=300)


    def change_apm(self,nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

if __name__=="__main__":
    app = App()
    app.mainloop()
